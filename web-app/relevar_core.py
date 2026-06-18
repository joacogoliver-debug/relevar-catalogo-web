"""
relevar_core.py — Motor de relevamiento de catálogos (sin CLI), para la app web.

Toma la URL de un canal de YouTube (Topic / Official Artist Channel / @handle),
enumera sus productos vía la YouTube Data API, opcionalmente enriquece con Spotify
(ISRC + UPC) y arma un Excel en memoria. Las claves se pasan como parámetros (la
app las toma de sus secrets), no se leen de archivos.

Función principal: relevar(url, yt_key, sp_id, sp_secret, progress) -> dict.
"""

import io
import json
import os
import re
import sys
import unicodedata
import urllib.parse
import urllib.request
import base64
import time
from collections import Counter
from concurrent.futures import ThreadPoolExecutor
from datetime import date
from difflib import SequenceMatcher

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

sys.path.insert(0, os.path.dirname(os.path.abspath(__file__)))
from distributors import (  # noqa: E402
    DIY_DISTRIBUTORS,
    DIFICIL_DISTRIBUTORS,
    DIFICIL_DISTRIBUTORS_EXACT,
    DISTRIBUTOR_BLACKLIST,
)

API = "https://www.googleapis.com/youtube/v3"


class RelevarError(Exception):
    """Error de negocio (canal no encontrado, clave inválida, etc.) para mostrar al usuario."""


# ============================================================
# Llamadas a la API
# ============================================================

def api_get(endpoint, params, key):
    params = dict(params)
    params["key"] = key
    url = f"{API}/{endpoint}?{urllib.parse.urlencode(params)}"
    req = urllib.request.Request(url, headers={"Accept": "application/json"})
    try:
        with urllib.request.urlopen(req, timeout=30) as resp:
            return json.loads(resp.read().decode("utf-8"))
    except urllib.error.HTTPError as e:
        body = e.read().decode("utf-8", "ignore")
        msg = body
        try:
            msg = json.loads(body)["error"]["message"]
        except Exception:
            pass
        raise RelevarError(f"ERROR de la API de YouTube ({e.code}): {msg}")


def resolve_channel(url, key):
    """Devuelve (channel_id, uploads_playlist_id, channel_title)."""
    url = url.strip()
    m = re.search(r"/channel/(UC[\w-]+)", url)
    handle = None
    if m:
        params = {"part": "snippet,contentDetails", "id": m.group(1)}
    else:
        hm = re.search(r"@([\w.\-]+)", url)
        if not hm:
            raise RelevarError("No pude extraer el canal de la URL. Usá una URL /channel/UC... o @handle.")
        handle = hm.group(1)
        params = {"part": "snippet,contentDetails", "forHandle": "@" + handle}

    data = api_get("channels", params, key)
    items = data.get("items") or []
    if not items:
        raise RelevarError(f"No se encontró el canal ({handle or url}). Revisá la URL.")
    ch = items[0]
    return (
        ch["id"],
        ch["contentDetails"]["relatedPlaylists"]["uploads"],
        ch["snippet"]["title"],
    )


def _is_topic_title(title):
    return title.strip().lower().endswith("- topic")


def channel_uploads_by_id(channel_id, key):
    """Devuelve (uploads_playlist_id, title) para un channel_id, o (None, None)."""
    data = api_get("channels", {"part": "snippet,contentDetails", "id": channel_id}, key)
    items = data.get("items") or []
    if not items:
        return None, None
    ch = items[0]
    return ch["contentDetails"]["relatedPlaylists"]["uploads"], ch["snippet"]["title"]


def find_topic_channel(artist_title, key):
    """Busca el canal '<artista> - Topic' (cuando te pasan un OAC). Devuelve
    channel_id o None. Cuesta 100 units de cuota (search.list)."""
    artist = re.sub(r"\s*-\s*Topic$", "", artist_title, flags=re.IGNORECASE).strip()
    data = api_get("search", {"part": "snippet", "type": "channel",
                              "q": f"{artist} - Topic", "maxResults": 5}, key)
    items = data.get("items", [])
    want = f"{artist.lower()} - topic"
    for it in items:  # match exacto primero
        if it["snippet"]["title"].strip().lower() == want:
            return it["snippet"]["channelId"]
    for it in items:  # match flexible: termina en "- topic" y contiene el artista
        t = it["snippet"]["title"].strip().lower()
        if t.endswith("- topic") and artist.lower() in t:
            return it["snippet"]["channelId"]
    return None


def list_video_ids(uploads_playlist, key):
    ids = []
    page = None
    while True:
        params = {"part": "contentDetails", "maxResults": 50, "playlistId": uploads_playlist}
        if page:
            params["pageToken"] = page
        data = api_get("playlistItems", params, key)
        for it in data.get("items", []):
            vid = it.get("contentDetails", {}).get("videoId")
            if vid:
                ids.append(vid)
        page = data.get("nextPageToken")
        if not page:
            break
    return ids


def fetch_videos(video_ids, key):
    out = []
    for i in range(0, len(video_ids), 50):
        batch = video_ids[i:i + 50]
        data = api_get("videos", {"part": "snippet,statistics,contentDetails", "id": ",".join(batch)}, key)
        out.extend(data.get("items", []))
    return out


_RE_ISO_DUR = re.compile(r"PT(?:(\d+)H)?(?:(\d+)M)?(?:(\d+)S)?")


def _iso_duration_to_seconds(s):
    if not s:
        return 0
    m = _RE_ISO_DUR.fullmatch(s)
    if not m:
        return 0
    h, mn, sec = (int(x) if x else 0 for x in m.groups())
    return h * 3600 + mn * 60 + sec


# ============================================================
# Parseo (distribuidora / álbum / año / sello)  — autocontenido
# ============================================================

_RE_PHONO_YEAR = re.compile(r"℗\s*(\d{4})\s+(.+)")
_RE_PHONO_LINE = re.compile(r"^\s*℗\s*(.+)$", re.MULTILINE)
_RE_LEADING_YEAR = re.compile(r"^\d{4}\s+")
_RE_RELEASED = re.compile(r"Released on:\s*(\d{4})-\d{2}-\d{2}")


def _normalize(s):
    return unicodedata.normalize("NFD", s).encode("ascii", "ignore").decode("ascii").lower()


def parse_description(desc):
    """Extrae distribuidor, álbum, año y sello de una descripción auto-generada."""
    res = {"distributor": None, "album": None, "release_year": None, "label": None}
    if not desc:
        return res

    first = desc.split("\n")[0].strip()
    prefix = "provided to youtube by "
    if first.lower().startswith(prefix):
        res["distributor"] = first[len(prefix):].strip()

    # Álbum: tercer bloque del formato auto-generado
    #   [0] Provided to YouTube by X / [1] Track · Artista / [2] Álbum
    blocks = [b.strip() for b in re.split(r"\n\s*\n", desc) if b.strip()]
    if len(blocks) >= 3 and blocks[0].lower().startswith(prefix):
        res["album"] = blocks[2].splitlines()[0].strip()

    # Año + sello
    m = _RE_PHONO_YEAR.search(desc)
    if m:
        res["release_year"] = int(m.group(1))
        res["label"] = m.group(2).strip()
    else:
        pm = _RE_PHONO_LINE.search(desc)
        if pm:
            res["label"] = _RE_LEADING_YEAR.sub("", pm.group(1).strip()).strip()
    if not res["release_year"]:
        rm = _RE_RELEASED.search(desc)
        if rm:
            res["release_year"] = int(rm.group(1))
    return res


def classify(distributor):
    if distributor is None:
        return "not_target"
    d = _normalize(distributor)
    if any(n in d for n in DISTRIBUTOR_BLACKLIST):
        return "blacklist"
    if any(n in d for n in DIY_DISTRIBUTORS):
        return "diy"
    if any(n in d for n in DIFICIL_DISTRIBUTORS) or d.strip() in DIFICIL_DISTRIBUTORS_EXACT:
        return "dificil"
    return "not_target"


def build_tracks(videos):
    tracks = []
    for v in videos:
        sn = v.get("snippet", {})
        st = v.get("statistics", {})
        cd = v.get("contentDetails", {})
        desc = sn.get("description") or ""
        meta = parse_description(desc)
        pub = (sn.get("publishedAt") or "")[:10]
        desc3 = "\n".join((desc.split("\n"))[:3]).strip()
        tracks.append({
            "video_id": v.get("id") or "",
            "track": sn.get("title") or "",
            "album": meta["album"] or "(single / sin álbum)",
            "distributor": meta["distributor"] or "(sin datos)",
            "category": classify(meta["distributor"]),
            "label": meta["label"] or "",
            "release_year": meta["release_year"] or "",
            "isrc": "",   # se completa por enriquecimiento (Spotify), si está disponible
            "upc": "",    # idem (a nivel álbum)
            "match": "",  # confianza del match con Spotify: alta / media / ""
            "duration_s": _iso_duration_to_seconds(cd.get("duration")),
            "views": int(st.get("viewCount", 0) or 0),
            "likes": int(st.get("likeCount", 0) or 0),
            "comments": int(st.get("commentCount", 0) or 0),
            "upload_date": pub,
            "desc3": desc3,
            "url": f"https://youtu.be/{v.get('id')}",
        })
    return tracks


# ============================================================
# Enriquecimiento de códigos (ISRC + UPC) — Deezer + respaldo MusicBrainz
# ============================================================
# Sin claves: Deezer y MusicBrainz son APIs públicas (adiós "se quedó sin
# créditos"). Deezer es la fuente principal (rápida y en paralelo: el ISRC viene
# en la búsqueda, el UPC por álbum). MusicBrainz es respaldo opcional para los
# tracks que Deezer no encuentre (lento: 1 pedido/seg, y cobertura despareja para
# artistas DIY → apagado por defecto). Match best-effort por título+artista+
# duración; ante la duda, código en blanco (mejor un hueco que un código errado).

DEEZER_API = "https://api.deezer.com"
MUSICBRAINZ_API = "https://musicbrainz.org/ws/2"
USER_AGENT = "RelevarCatalogo/1.1 (https://mojo-latam)"
# Deezer permite ~50 pedidos/5s por IP. Con 6 hilos + reintento ante quota,
# quedamos rápidos sin que nos corte. (MusicBrainz va aparte, secuencial.)
CODES_WORKERS = 6


def _deezer_json(path):
    """GET a Deezer con reintento. Deezer señala el límite de tasa con un JSON
    de error a HTTP 200 ({"error": {...}}), así que lo detectamos y reintentamos."""
    url = f"{DEEZER_API}/{path}"
    for _ in range(6):
        data = _http_json(url)
        if isinstance(data, dict) and data.get("error"):
            time.sleep(1.5)
            continue
        return data
    return None

# Sufijos de YouTube que ensucian el match (no están en el catálogo del DSP).
_RE_TITLE_NOISE = re.compile(
    r"\((?:[^)]*?(?:video|oficial|official|audio|en vivo|live|lyric|letra|"
    r"visualizer|remaster|hd|4k|cover)[^)]*?)\)|\[[^\]]*\]", re.IGNORECASE)


def _clean_title(title):
    t = _RE_TITLE_NOISE.sub("", title or "")
    return re.sub(r"\s+", " ", t).strip(" -–·")


def _http_json(url, headers=None, retries=3):
    req = urllib.request.Request(url, headers=headers or {})
    for _ in range(retries):
        try:
            with urllib.request.urlopen(req, timeout=25) as r:
                return json.loads(r.read().decode("utf-8"))
        except urllib.error.HTTPError as e:
            if e.code in (429, 500, 502, 503):
                time.sleep(int(e.headers.get("Retry-After", "1")) + 1)
                continue
            return None
        except Exception:
            time.sleep(1)
            continue
    return None


def _match_score(yt_title, yt_artist, yt_dur, cand_title, cand_artist, cand_dur):
    ratio = SequenceMatcher(None, _normalize(_clean_title(yt_title)),
                            _normalize(cand_title or "")).ratio()
    na, ns = _normalize(yt_artist), _normalize(cand_artist or "")
    artist_ok = bool(na) and (na in ns or ns in na or
                              SequenceMatcher(None, na, ns).ratio() >= 0.6)
    dur_close = None
    if yt_dur and cand_dur:
        dur_close = abs(yt_dur - cand_dur) <= 4
    score = ratio + (0.10 if artist_ok else 0) + (0.10 if dur_close else 0)
    return score, ratio, artist_ok, dur_close


def _confidence(ratio, artist_ok, dur_close):
    if ratio >= 0.87 and artist_ok and dur_close is not False:
        return "alta"
    if ratio >= 0.72 and artist_ok:
        return "media"
    return ""


# ---- Deezer (principal, sin clave) ----
def deezer_match(t, artist):
    """Devuelve (isrc, album_id, confianza) o ('', None, '')."""
    title = _clean_title(t["track"])
    q = urllib.parse.quote(f'track:"{title}" artist:"{artist}"')
    items = (_deezer_json(f"search?q={q}&limit=5") or {}).get("data") or []
    if not items:  # reintento con búsqueda libre (el filtro estricto a veces no matchea)
        q2 = urllib.parse.quote(f"{title} {artist}")
        items = (_deezer_json(f"search?q={q2}&limit=5") or {}).get("data") or []
    if not items:
        return "", None, ""
    best, best_meta = None, None
    for c in items:
        sc = _match_score(t["track"], artist, t.get("duration_s", 0),
                          c.get("title", ""), (c.get("artist") or {}).get("name", ""),
                          c.get("duration", 0))
        if best is None or sc[0] > best[0]:
            best, best_meta = sc, c
    conf = _confidence(best[1], best[2], best[3])
    if not conf:
        return "", None, ""
    isrc = best_meta.get("isrc") or ""
    if not isrc:  # algunos resultados no traen ISRC en la búsqueda → pedir el track
        isrc = (_deezer_json(f"track/{best_meta.get('id')}") or {}).get("isrc") or ""
    return isrc, (best_meta.get("album") or {}).get("id"), conf


def deezer_album_upcs(album_ids):
    ids = [a for a in album_ids if a]
    if not ids:
        return {}

    def work(aid):
        return aid, (_deezer_json(f"album/{aid}") or {}).get("upc", "") or ""

    out = {}
    with ThreadPoolExecutor(max_workers=min(CODES_WORKERS, len(ids))) as ex:
        for aid, upc in ex.map(work, ids):
            out[aid] = upc
    return out


# ---- MusicBrainz (respaldo opcional; límite 1 pedido/seg) ----
def musicbrainz_isrc(t, artist):
    """ISRC desde MusicBrainz. 2 pedidos (search + lookup). Devuelve '' si no hay."""
    q = urllib.parse.quote(f'recording:"{_clean_title(t["track"])}" AND artist:"{artist}"')
    data = _http_json(f"{MUSICBRAINZ_API}/recording?query={q}&fmt=json&limit=5",
                      headers={"User-Agent": USER_AGENT})
    for r in (data or {}).get("recordings", []) or []:
        ac = " ".join(a.get("name", "") for a in (r.get("artist-credit") or [])
                      if isinstance(a, dict))
        dur = round((r.get("length") or 0) / 1000)
        sc = _match_score(t["track"], artist, t.get("duration_s", 0),
                          r.get("title", ""), ac, dur)
        if _confidence(sc[1], sc[2], sc[3]):
            time.sleep(1.1)  # respetar el límite de MusicBrainz entre los 2 pedidos
            look = _http_json(f"{MUSICBRAINZ_API}/recording/{r['id']}?fmt=json&inc=isrcs",
                              headers={"User-Agent": USER_AGENT})
            isrcs = (look or {}).get("isrcs") or []
            return isrcs[0] if isrcs else ""
    return ""


def enrich_with_codes(tracks, artist, log=print, use_musicbrainz=False):
    """Completa isrc/upc/match en los tracks. Deezer principal + MB opcional."""
    n = len(tracks)

    # 1) Deezer en paralelo (ISRC + album_id por track).
    def work(t):
        return (t,) + deezer_match(t, artist)

    album_ids = {}
    matched = 0
    if n:
        with ThreadPoolExecutor(max_workers=min(CODES_WORKERS, n)) as ex:
            for t, isrc, album_id, conf in ex.map(work, tracks):
                if conf:
                    t["isrc"], t["match"] = isrc, conf
                    if album_id:
                        album_ids.setdefault(album_id, []).append(t)
                    matched += 1

    # 2) UPC por álbum (Deezer, en paralelo).
    log(f"  Deezer: {matched}/{n} matcheados · UPC de {len(album_ids)} álbumes…")
    upcs = deezer_album_upcs(list(album_ids.keys()))
    for aid, ts in album_ids.items():
        for t in ts:
            t["upc"] = upcs.get(aid, "")

    # 3) MusicBrainz: respaldo SÓLO para los que quedaron sin ISRC (secuencial, lento).
    if use_musicbrainz:
        pendientes = [t for t in tracks if not t["isrc"]]
        if pendientes:
            log(f"  MusicBrainz (respaldo): {len(pendientes)} sin ISRC…")
            for t in pendientes:
                isrc = musicbrainz_isrc(t, artist)
                if isrc:
                    t["isrc"] = isrc
                    t["match"] = t["match"] or "media"
                time.sleep(1.1)  # 1 pedido/seg

    isrc_n = sum(1 for t in tracks if t["isrc"])
    upc_n = sum(1 for t in tracks if t["upc"])
    return {"matched": matched, "isrc": isrc_n, "upc": upc_n, "source": "Deezer"}


# ============================================================
# Excel  (Resumen tipo dashboard + hoja de detalle)
# ============================================================

NAVY = "1A1A2E"
GRAY = "6B7280"
PANEL = "F7F7F5"
WHITE = "FFFFFF"
RED = "CC0000"
RED_BRIGHT = "FF0000"
RED_BG1 = "FFF8F8"
RED_BG2 = "FFF0F0"

THIN = Side(style="thin", color="E5E7EB")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
NUMFMT = "#,##0"


def _f(size=10, bold=False, color=NAVY):
    return Font(name="Calibri", size=size, bold=bold, color=color)


def _fill(hexcolor):
    return PatternFill("solid", fgColor=hexcolor)


def _set(ws, coord, value, font=None, fill=None, align="left", numfmt=None, wrap=False):
    c = ws[coord]
    c.value = value
    if font:
        c.font = font
    if fill:
        c.fill = fill
    c.alignment = Alignment(horizontal=align, vertical="center", wrap_text=wrap)
    if numfmt:
        c.number_format = numfmt
    return c


def _fill_range(ws, rng, fill):
    """Pinta el fondo de todas las celdas de un rango combinado."""
    for row in ws[rng]:
        for c in row:
            c.fill = fill


def _aggregate_distributors(tracks):
    agg = {}
    for t in tracks:
        d = agg.setdefault(t["distributor"], {"videos": 0, "views": 0, "top": 0, "top_title": ""})
        d["videos"] += 1
        d["views"] += t["views"]
        if t["views"] > d["top"]:
            d["top"] = t["views"]
            d["top_title"] = t["track"]
    return agg


def build_resumen(wb, tracks, artist):
    ws = wb.active
    ws.title = "Resumen"
    ws.sheet_view.showGridLines = False
    widths = {"A": 2, "B": 18, "C": 22, "D": 18, "E": 22, "F": 18, "G": 22, "H": 2}
    for col, w in widths.items():
        ws.column_dimensions[col].width = w

    total_videos = len(tracks)
    total_views = sum(t["views"] for t in tracks)
    years = sorted({t["release_year"] for t in tracks if t["release_year"]})
    period = f"{years[0]} – {years[-1]}" if years else "s/d"
    gen = date.today().strftime("%d/%m/%Y")

    agg = _aggregate_distributors(tracks)
    # Distribuidora protagonista = la de más vistas totales
    spot = max(agg.items(), key=lambda kv: kv[1]["views"]) if agg else (None, None)
    spot_name, spot_data = spot

    # --- Título ---
    ws.merge_cells("B2:G2")
    _set(ws, "B2", f"{artist.upper()} — ANÁLISIS DE CATÁLOGO (TOPIC)", _f(18, True, NAVY), _fill(PANEL))
    ws.merge_cells("B3:G3")
    subtitle = f"Dataset: {total_videos} videos  ·  Período: {period}  ·  Generado: {gen}"
    isrc_n = sum(1 for t in tracks if t.get("isrc"))
    if isrc_n:
        upc_n = sum(1 for t in tracks if t.get("upc"))
        subtitle += (f"  ·  ISRC: {isrc_n}/{total_videos}"
                     f"  ·  UPC: {upc_n}/{total_videos} (vía Spotify)")
    _set(ws, "B3", subtitle, _f(9, False, GRAY), _fill(PANEL))
    _fill_range(ws, "B2:G2", _fill(PANEL))
    _fill_range(ws, "B3:G3", _fill(PANEL))
    ws.row_dimensions[2].height = 26

    # --- KPI cards (filas 6-10) ---
    for rng in ("B6:D10", "E6:G10"):
        _fill_range(ws, rng, _fill(WHITE))
    for a, b in (("B6", "D6"), ("B7", "D7"), ("B8", "D8"), ("B9", "D9"), ("B10", "D10"),
                 ("E6", "G6"), ("E7", "G7"), ("E8", "G8"), ("E9", "G9"), ("E10", "G10")):
        ws.merge_cells(f"{a}:{b}")
    _set(ws, "B7", "TOTAL VIDEOS", _f(8, False, GRAY), _fill(WHITE))
    _set(ws, "B8", total_videos, _f(20, True, NAVY), _fill(WHITE), numfmt=NUMFMT)
    _set(ws, "B9", "Productos en el catálogo", _f(9, False, GRAY), _fill(WHITE))
    _set(ws, "E7", "TOTAL REPRODUCCIONES", _f(8, False, GRAY), _fill(WHITE))
    _set(ws, "E8", total_views, _f(20, True, NAVY), _fill(WHITE), numfmt=NUMFMT)
    _set(ws, "E9", "Vistas acumuladas", _f(9, False, GRAY), _fill(WHITE))

    # --- Spotlight distribuidora (filas 12-16) ---
    if spot_name:
        for rng in ("B12:C16", "D12:E16", "F12:G16"):
            _fill_range(ws, rng, _fill(RED_BG1))
        for a, b in (("B12", "C12"), ("B13", "C13"), ("B14", "C14"), ("B15", "C15"), ("B16", "C16"),
                     ("D12", "E12"), ("D13", "E13"), ("D14", "E14"), ("D15", "E15"), ("D16", "E16"),
                     ("F12", "G12"), ("F13", "G13"), ("F14", "G14"), ("F15", "G15"), ("F16", "G16")):
            ws.merge_cells(f"{a}:{b}")
        sv = spot_data["videos"]
        svw = spot_data["views"]
        _set(ws, "B13", f"VIDEOS — {spot_name.upper()[:22]}", _f(8, False, RED), _fill(RED_BG1))
        _set(ws, "B14", sv, _f(18, True, RED_BRIGHT), _fill(RED_BG1), numfmt=NUMFMT)
        _set(ws, "B15", f"{sv/total_videos*100:.1f}% del total", _f(9, False, RED), _fill(RED_BG1))
        _set(ws, "D13", f"VISTAS — {spot_name.upper()[:22]}", _f(8, False, RED), _fill(RED_BG1))
        _set(ws, "D14", svw, _f(18, True, RED_BRIGHT), _fill(RED_BG1), numfmt=NUMFMT)
        _set(ws, "D15", f"{(svw/total_views*100 if total_views else 0):.1f}% del total", _f(9, False, RED), _fill(RED_BG1))
        _set(ws, "F13", f"TOP VIDEO — {spot_name.upper()[:18]}", _f(8, False, RED), _fill(RED_BG1))
        _set(ws, "F14", spot_data["top"], _f(18, True, RED_BRIGHT), _fill(RED_BG1), numfmt=NUMFMT)
        _set(ws, "F15", spot_data["top_title"][:40], _f(9, False, RED), _fill(RED_BG1), wrap=True)

    # --- TOP 10 videos ---
    ws.merge_cells("B18:G18")
    _set(ws, "B18", "TOP 10 VIDEOS POR REPRODUCCIONES", _f(10, True, NAVY), _fill(PANEL))
    _fill_range(ws, "B18:G18", _fill(PANEL))
    ws.merge_cells("C19:D19")
    _set(ws, "B19", "#", _f(9, True, WHITE), _fill(NAVY), align="center")
    _set(ws, "C19", "Título", _f(9, True, WHITE), _fill(NAVY))
    _set(ws, "E19", "Distribuidora", _f(9, True, WHITE), _fill(NAVY))
    _set(ws, "F19", "Vistas", _f(9, True, WHITE), _fill(NAVY), align="right")
    _set(ws, "G19", "Año", _f(9, True, WHITE), _fill(NAVY), align="center")
    _fill_range(ws, "C19:D19", _fill(NAVY))
    top10 = sorted(tracks, key=lambda x: x["views"], reverse=True)[:10]
    for i, t in enumerate(top10):
        r = 20 + i
        is_top = (i == 0)
        bg = _fill(RED_BG2) if is_top else _fill(WHITE)
        fc = RED_BRIGHT if is_top else NAVY
        ws.merge_cells(f"C{r}:D{r}")
        _fill_range(ws, f"C{r}:D{r}", bg)
        _set(ws, f"B{r}", i + 1, _f(9, is_top, fc), bg, align="center")
        _set(ws, f"C{r}", t["track"][:48], _f(9, False, fc), bg)
        _set(ws, f"E{r}", t["distributor"], _f(9, False, fc), bg)
        _set(ws, f"F{r}", t["views"], _f(9, True, fc), bg, align="right", numfmt=NUMFMT)
        _set(ws, f"G{r}", str(t["release_year"]) if t["release_year"] else "", _f(9, False, fc), bg, align="center")

    # --- Distribución por distribuidora ---
    ws.merge_cells("B32:G32")
    _set(ws, "B32", "DISTRIBUCIÓN POR DISTRIBUIDORA", _f(10, True, NAVY), _fill(PANEL))
    _fill_range(ws, "B32:G32", _fill(PANEL))
    ws.merge_cells("E33:G33")
    _set(ws, "B33", "Distribuidora", _f(9, True, WHITE), _fill(NAVY))
    _set(ws, "C33", "Videos", _f(9, True, WHITE), _fill(NAVY), align="right")
    _set(ws, "D33", "% Total", _f(9, True, WHITE), _fill(NAVY), align="right")
    _set(ws, "E33", "Vistas totales", _f(9, True, WHITE), _fill(NAVY), align="right")
    _fill_range(ws, "E33:G33", _fill(NAVY))
    ordered = sorted(agg.items(), key=lambda kv: kv[1]["videos"], reverse=True)
    for i, (name, d) in enumerate(ordered):
        r = 34 + i
        ws.merge_cells(f"E{r}:G{r}")
        _fill_range(ws, f"E{r}:G{r}", _fill(WHITE))
        _set(ws, f"B{r}", name, _f(9, False, NAVY), _fill(WHITE))
        _set(ws, f"C{r}", d["videos"], _f(9, False, NAVY), _fill(WHITE), align="right", numfmt=NUMFMT)
        _set(ws, f"D{r}", d["videos"] / total_videos, _f(9, False, NAVY), _fill(WHITE), align="right", numfmt="0.0%")
        _set(ws, f"E{r}", d["views"], _f(9, False, NAVY), _fill(WHITE), align="right", numfmt=NUMFMT)


def build_detail(wb, tracks, artist):
    title = (artist + " Topic")[:31]
    ws = wb.create_sheet(title)
    cols = ["VIDEO_ID", "TITLE", "PUBLISHED_AT", "REPRODUCCIONES", "DISTRIBUIDORA",
            "SELLO", "ISRC", "UPC", "MATCH", "DESCRIPTION (3 LÍNEAS)", "URL"]
    for ci, name in enumerate(cols, 1):
        c = ws.cell(row=1, column=ci, value=name)
        c.font = _f(10, True, WHITE)
        c.fill = _fill(NAVY)
        c.alignment = Alignment(horizontal="left", vertical="center")
    for t in sorted(tracks, key=lambda x: x["views"], reverse=True):
        ws.append([t["video_id"], t["track"], t["upload_date"], t["views"],
                   t["distributor"], t["label"], t["isrc"], t["upc"], t.get("match", ""),
                   t["desc3"], t["url"]])
    for col, w in zip("ABCDEFGHIJK", (14, 40, 13, 15, 24, 22, 15, 15, 9, 46, 28)):
        ws.column_dimensions[col].width = w
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
        row[3].number_format = NUMFMT          # REPRODUCCIONES
        row[9].alignment = Alignment(wrap_text=True, vertical="top")  # DESCRIPTION
    ws.freeze_panes = "A2"


def build_workbook(tracks, artist):
    wb = openpyxl.Workbook()
    build_resumen(wb, tracks, artist)
    build_detail(wb, tracks, artist)
    return wb


def workbook_bytes(tracks, artist):
    """Arma el Excel y lo devuelve como bytes (para descargar desde la web)."""
    buf = io.BytesIO()
    build_workbook(tracks, artist).save(buf)
    return buf.getvalue()


def slugify(name):
    s = re.sub(r"[^\w\s-]", "", name, flags=re.UNICODE).strip()
    return re.sub(r"\s+", "_", s) or "Artista"


# ============================================================
# Orquestador (lo llama la app web)
# ============================================================

def relevar(url, yt_key, with_codes=True, progress=None, use_musicbrainz=False):
    """Releva el catálogo completo de un canal.

    Si le pasás un OAC (u otro canal que no es Topic), busca automáticamente su
    canal '<artista> - Topic' y usa ese. Siempre se queda SOLO con los lanzamientos
    auto-generados (los que tienen "Provided to YouTube by"), descartando videos
    comunes (vlogs, vivos, MVs, etc.).

    with_codes — buscar ISRC/UPC (Deezer; sin clave). use_musicbrainz — respaldo
    lento opcional. progress(msg, frac) — callback de avance (0.0–1.0).
    Lanza RelevarError ante problemas mostrables al usuario.
    """
    def step(msg, frac):
        if progress:
            progress(msg, frac)

    if not yt_key:
        raise RelevarError("Falta la API key de YouTube en el servidor.")

    step("Resolviendo canal…", 0.05)
    _ch_id, uploads, title = resolve_channel(url, yt_key)

    # Si NO es un canal Topic (ej. un OAC), buscamos su canal "- Topic" automáticamente.
    via_topic = False
    if not _is_topic_title(title):
        step("Es un OAC: buscando su canal Topic…", 0.10)
        topic_id = find_topic_channel(title, yt_key)
        if topic_id:
            t_uploads, t_title = channel_uploads_by_id(topic_id, yt_key)
            if t_uploads:
                uploads, title, via_topic = t_uploads, t_title, True

    step("Listando productos…", 0.15)
    vids = list_video_ids(uploads, yt_key)
    if not vids:
        raise RelevarError("El canal no tiene productos para relevar.")

    step(f"Bajando metadata de {len(vids)} productos…", 0.30)
    videos = fetch_videos(vids, yt_key)

    # Solo lanzamientos: nos quedamos con los auto-generados (tienen distribuidora
    # parseada de "Provided to YouTube by"); descartamos videos comunes.
    tracks = [t for t in build_tracks(videos) if t["distributor"] != "(sin datos)"]
    if not tracks:
        raise RelevarError(
            "No encontré lanzamientos auto-generados (Topic) en este canal. "
            "Si es un OAC, probá pegar el link de su canal “<artista> - Topic”.")

    artist = re.sub(r"\s*-\s*Topic$", "", title).strip()

    codes_stats = None
    if with_codes:
        step("Buscando códigos ISRC y UPC (Deezer)…", 0.55)
        codes_stats = enrich_with_codes(
            tracks, artist, log=lambda m: step(m, 0.75), use_musicbrainz=use_musicbrainz)

    step("Armando el Excel…", 0.95)
    units = 1 + 2 * ((len(vids) + 49) // 50)
    return {
        "artist": artist,
        "channel_title": title,
        "tracks": tracks,
        "distribs": _aggregate_distributors(tracks),
        "total_views": sum(t["views"] for t in tracks),
        "units": units,
        "codes": codes_stats,
        "via_topic": via_topic,
    }
